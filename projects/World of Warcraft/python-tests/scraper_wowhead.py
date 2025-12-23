import os
import re
import time
import random
import pandas as pd
from openpyxl import Workbook, load_workbook
from scraper_wiki import pobierz_soup


def wyciagnij_patch(soup):
    for s in soup.find_all("script"):
        t = s.get_text(" ", strip=True)
        if "Added in patch" not in t:
            continue
        m = re.search(
            r'Added in patch\s*\[acronym=\\?"[^"]*\\?"\]([0-9]+\.[0-9]+\.[0-9]+)\[\\?/acronym\]',
            t
        )
        if m:
            return m.group(1)
    return ""

def normalize_cell(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if pd.isna(v):
        return None
    return v

def buduj_mapping_01():
    """
    Tworzy nowego excela, który jest rozbudowany o nazwę linii fabularnej oraz w którym patchu dodano misję.
    Funkcja bazuje na excelu "surowym", który znajduje się w folderze "surowe" -> plik "wowhead_id_kraina_dodatek".
    """
    raw_path = r"D:\MyProjects_4Fun\projects\World of Warcraft\excel-mappingi\surowe\wowhead_id_kraina_dodatek.xlsx"
    out_path = r"D:\MyProjects_4Fun\projects\World of Warcraft\excel-mappingi\mapping_01.xlsx"

    input_sheet = "prawie_gotowe_dane"
    output_sheet = "mapping_01"
    url_col = "MISJA_URL_WOWHEAD"

    batch_size = 10

    df_raw = pd.read_excel(raw_path, sheet_name=input_sheet)
    print(f"Odczytano {len(df_raw)} wierszy z pliku surowego")

    if url_col not in df_raw.columns:
        raise ValueError(f"Brak kolumny {url_col}")

    df_raw[url_col] = df_raw[url_col].astype(str).str.strip()
    df_raw = df_raw[df_raw[url_col].notna() & (df_raw[url_col] != "")].copy()

    headers = list(df_raw.columns) + ["storyline", "patch"]

    if not os.path.exists(out_path):
        wb = Workbook()
        ws = wb.active
        ws.title = output_sheet
        ws.append(headers)
        wb.save(out_path)
        print("Utworzono nowy plik mapping_01.xlsx")

    df_out = pd.read_excel(out_path, sheet_name=output_sheet)
    if url_col in df_out.columns:
        existing_urls = set(
            df_out[url_col]
            .dropna()
            .astype(str)
            .str.strip()
            .tolist()
        )
    else:
        existing_urls = set()

    print(f"W pliku wynikowym jest już {len(existing_urls)} URL-i")

    df_new = df_raw[~df_raw[url_col].isin(existing_urls)].copy()
    print(f"Nowych questów do pobrania: {len(df_new)}")

    if df_new.empty:
        print("Nic nowego do zrobienia")
        return

    wb = load_workbook(out_path)
    ws = wb[output_sheet]

    bufor = []
    dopisane = 0
    bledy = 0

    for idx, row in enumerate(df_new.itertuples(index=False), start=1):
        row_dict = row._asdict()
        link = row_dict[url_col]

        print(f"[{idx}/{len(df_new)}] Scrapuję: {link}")
        soup = pobierz_soup(link, parser="lxml")

        if soup is None:
            bledy += 1
            print("  Błąd pobierania strony")
            continue

        element = soup.select_one(".quick-facts-storyline-title")
        storyline = element.get_text().strip() if element else ""

        if not storyline:
            print("  Brak storyline — zapisuję pusty")

        patch = wyciagnij_patch(soup)

        if not patch:
            print("  Brak patcha")

        row_dict["storyline"] = storyline
        row_dict["patch"] = patch

        out_row = [normalize_cell(row_dict.get(col)) for col in headers]
        bufor.append(out_row)

        print(f"  OK | storyline='{storyline}' | patch='{patch}'")

        if len(bufor) >= batch_size:
            start_row = ws.max_row + 1
            for r in bufor:
                ws.append(r)
            wb.save(out_path)
            dopisane += len(bufor)
            print(f"Zapisano paczkę {len(bufor)} wierszy od wiersza {start_row}")
            bufor = []

        time.sleep(random.uniform(1.3, 1.9))

    if bufor:
        start_row = ws.max_row + 1
        for r in bufor:
            ws.append(r)
        wb.save(out_path)
        dopisane += len(bufor)
        print(f"Zapisano ostatnią paczkę {len(bufor)} wierszy od wiersza {start_row}")

    print(f"Koniec. Dopisano: {dopisane}, błędy pobierania: {bledy}")