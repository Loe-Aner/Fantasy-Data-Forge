import asyncio
import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from sqlalchemy import text
from scraper_wiki_async import WoWScraperService

# --- IMPORT O KTÓRY PROSIŁEŚ ---
from moduly.db_core import utworz_engine_do_db 

def wyciagnij_patch_final(soup: BeautifulSoup) -> str:
    element_tekstowy = soup.find(string=re.compile("Added in patch"))
    
    if element_tekstowy:
        rodzic = element_tekstowy.parent
        caly_tekst = rodzic.get_text(separator=" ", strip=True)
        
        dopasowanie = re.search(r"(\d+\.\d+\.\d+)", caly_tekst)
        if dopasowanie:
            return dopasowanie.group(1)
            
    return ""

def wyciagnij_storyline_final(soup: BeautifulSoup) -> str:
    element_tytul = soup.select_one(".quick-facts-storyline-title")
    if element_tytul:
        return element_tytul.get_text(strip=True)

    naglowek = soup.find("th", string=re.compile("Storyline"))
    if naglowek:
        wiersz_naglowka = naglowek.find_parent("tr")
        if wiersz_naglowka:
            wiersz_danych = wiersz_naglowka.find_next_sibling("tr")
            if wiersz_danych:
                kontener = wiersz_danych.select_one(".quick-facts-storyline")
                if kontener:
                    link = kontener.find("a")
                    if link:
                        return link.get_text(strip=True)
    
    return ""

def wyciagnij_kraine_ze_skryptu(html_text: str) -> str:
    if not html_text:
        return ""
        
    dopasowanie = re.search(r'"zone":"([^"]+)"', html_text)
    
    if dopasowanie:
        return dopasowanie.group(1)
        
    return ""

def okresl_dodatek_na_podstawie_patcha(patch: str) -> str:
    if not patch:
        return ""
    
    glowna_wersja = patch.split('.')[0]
    
    mapping = {
        "1": "Vanilla / Classic",
        "2": "The Burning Crusade",
        "3": "Wrath of the Lich King",
        "4": "Cataclysm",
        "5": "Mists of Pandaria",
        "6": "Warlords of Draenor",
        "7": "Legion",
        "8": "Battle for Azeroth",
        "9": "Shadowlands",
        "10": "Dragonflight",
        "11": "The War Within",
        "12": "Midnight",
        "13": "The Last Titan"
    }
    
    return mapping.get(glowna_wersja, "Brak Dodatku")

def parsuj_wowhead_html(html: str, url: str) -> tuple[str, str, str, str, str, str]:
    if not html:
        return url, "", "", "", "", "Brak HTML"

    try:
        soup = BeautifulSoup(html, "lxml")
    except Exception:
        soup = BeautifulSoup(html, "html.parser")

    linia_fabularna = wyciagnij_storyline_final(soup)
    patch = wyciagnij_patch_final(soup)
    
    kraina = wyciagnij_kraine_ze_skryptu(html)
    dodatek = okresl_dodatek_na_podstawie_patcha(patch)
    
    return url, linia_fabularna, patch, dodatek, kraina, ""

class WowheadScraper(WoWScraperService):
    async def process_url(self, url: str):
        html = await self._fetch_html(url)
        
        if not html:
            return url, "", "", "", "", "Błąd pobierania (HTTP)"

        loop = asyncio.get_running_loop()
        wynik = await loop.run_in_executor(None, parsuj_wowhead_html, html, url)
        
        return wynik

def normalize_cell(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if pd.isna(v):
        return None
    return v

def zapisz_excel_w_tle(wb: Workbook, path: str):
    wb.save(path)

async def buduj_mapping_01_async():
    out_path = r"D:\MyProjects_4Fun\projects\World of Warcraft\excel-mappingi\mapping_01.xlsx"
    output_sheet = "mapping_01"
    url_col = "MISJA_URL_WOWHEAD"

    FINAL_HEADERS = [
        "KRAINA_EN",
        "MISJA_ID_Z_GRY",
        "MISJA_TYTUL_EN",
        "DODATEK_EN",
        "MISJA_URL_WOWHEAD",
        "NAZWA_LINII_FABULARNEJ_EN",
        "DODANO_W_PATCHU"
    ]

    MAX_CONCURRENCY = 8
    BATCH_SIZE = 48

    print("Łączę z bazą danych SQL...")
    
    zapytanie_sql = """
        SELECT 
            KRAINA_EN,
            MISJA_ID_Z_GRY,
            MISJA_TYTUL_EN,
            DODATEK_EN,
            MISJA_URL_WOWHEAD
        FROM dbo.MISJE
        WHERE MISJA_URL_WOWHEAD IS NOT NULL 
          AND MISJA_URL_WOWHEAD != ''
    """

    silnik = utworz_engine_do_db()
    with silnik.connect() as polaczenie:
        df_raw = pd.read_sql(text(zapytanie_sql), polaczenie)

    print(f"Pobrano {len(df_raw)} wierszy z bazy danych.")

    df_raw[url_col] = df_raw[url_col].astype(str).str.strip()
    
    if not os.path.exists(out_path):
        wb = Workbook()
        ws = wb.active
        ws.title = output_sheet
        ws.append(FINAL_HEADERS)
        wb.save(out_path)
        print("Utworzono nowy plik mapping_01.xlsx")
    else:
        try:
            pd.read_excel(out_path, sheet_name=output_sheet, nrows=0)
        except Exception as e:
            print(f"Błąd odczytu pliku wyjściowego: {e}")
            return

    df_out = pd.read_excel(out_path, sheet_name=output_sheet)
    
    if url_col in df_out.columns:
        existing_urls = set(df_out[url_col].dropna().astype(str).str.strip().tolist())
    else:
        existing_urls = set()

    df_new = df_raw[~df_raw[url_col].isin(existing_urls)].copy()
    print(f"Postęp: {len(existing_urls)} zrobionych. Nowych do pobrania: {len(df_new)}")

    if df_new.empty:
        print("Wszystko zrobione! Idź pograć w WoW-a.")
        return

    row_by_url = {
        str(row[url_col]).strip(): row.to_dict()
        for _, row in df_new.iterrows()
    }
    
    urls_to_process = list(row_by_url.keys())

    wb = load_workbook(out_path)
    ws = wb[output_sheet]

    scraper = WowheadScraper(concurrency=MAX_CONCURRENCY)
    loop = asyncio.get_running_loop()

    dopisane = 0
    bledy = 0

    try:
        print(f"Start scrapowania {len(urls_to_process)} misji...")
        
        for start in range(0, len(urls_to_process), BATCH_SIZE):
            batch_urls = urls_to_process[start : start + BATCH_SIZE]
            print(f"\nBatch {start + 1}-{start + len(batch_urls)} / {len(urls_to_process)}")

            tasks = [scraper.process_url(u) for u in batch_urls]
            wyniki = await asyncio.gather(*tasks)

            bufor = []
            
            for link, storyline, patch, dodatek_wh, kraina_wh, err in wyniki:
                if err:
                    bledy += 1
                    print(f" [FAIL] {link} -> {err}")
                    continue
                
                row_data = row_by_url.get(link)
                if not row_data:
                    continue

                row_data["NAZWA_LINII_FABULARNEJ_EN"] = storyline
                row_data["DODANO_W_PATCHU"] = patch
                
                if dodatek_wh:
                    row_data["DODATEK_EN"] = dodatek_wh
                
                if kraina_wh:
                    row_data["KRAINA_EN"] = kraina_wh

                excel_row = []
                for header in FINAL_HEADERS:
                    val = row_data.get(header)
                    excel_row.append(normalize_cell(val))

                bufor.append(excel_row)
                print(f" [OK] {link} (P: {patch} | D: {dodatek_wh} | K: {kraina_wh})")

            if bufor:
                for r in bufor:
                    ws.append(r)

                print(" -> Zapisuję batch do pliku...")
                await loop.run_in_executor(None, zapisz_excel_w_tle, wb, out_path)
                dopisane += len(bufor)

    finally:
        await scraper.close()

    print(f"\nPodsumowanie: Sukces: {dopisane}, Błędy: {bledy}")

def buduj_mapping_01():
    asyncio.run(buduj_mapping_01_async())

if __name__ == "__main__":
    buduj_mapping_01()